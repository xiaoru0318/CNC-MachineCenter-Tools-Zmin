# -*- coding: utf-8 -*-
"""
Created on Wed Jun 29 21:25:55 2022

@author: XiaoRu0318
"""


from re import compile
import openpyxl 
import os
from openpyxl.styles import Font,Side,Border,colors,Alignment #设置单元格格式
import time

def file_searcher():    #搜索当前文件夹的数据文件
    path_main = os.getcwd() #获取当前路径
    z_1 = 'nc'    #excel文件类型
    z_2 = 'MIN'

    file_list = os.listdir(path_main)   #获取当前文件夹中的文件
    choose_list = []     
    ex_1 = compile(r'(\.'+z_1+'$)')  #筛选符合的文件名称及格式
    ex_2 = compile(r'(\.'+z_2+'$)')

    for i in file_list:                    
        if ex_1.search(i) or ex_2.search(i):
            choose_list.append(i)
    return choose_list   

if __name__ == "__main__":  #主函数  
    choose_list = file_searcher()

result = []

for h in choose_list:
    #print(h)      
    f = open(h,'r')        
    something =  f.readlines()      
    T_checker = compile(r'T(\d)+(.*)') #查找刀号，\d为数字，+为匹配一次或者多次
    Z_checker = compile(r'Z(-)?(\d)+(\.)?(\d)*') #查找Z的相关数据
    T_temp = 0
    Z_temp = 0   
    
    for x in something:
        T_get = T_checker.search(x)
        if T_get != None :
            T_temp = T_get.group()
            #print(T_temp)
        Z_get = Z_checker.search(x)
        if Z_get != None :
            Z_temp = Z_get.group()
            #print(Z_temp)
            result.append([T_temp,Z_temp])           

    f.close()
 
def print_array(arr):
    obj:dict = {} #暂时存储的空字典
    for item in arr:
        if not obj.get(item[0]): #从字典中找到key，如果没有，就赋值给空集里
            obj[item[0]] = item[1]
        elif float(obj.get(item[0])[1:]) > float(item[1][1:]): #从字典中取value，如果数值小就赋值给空集里
            obj[item[0]] = item[1]
               
    return [[x,y] for x,y in obj.items()]
 
if __name__ == '__main__':
    final = print_array(result)
   
for k in final:
    if k == [0,'Z1000.']:
        final.remove(k)
        #print(final)
       
final.sort(key=lambda final:(final[0][3],int(final[0][1]),int(final[0][2])),reverse=False) #结果排序
    #print(final)

wb = openpyxl.Workbook()  # 实例化一个工作簿对象
ws = wb['Sheet'] #激活工作表单
#ws=wb.active()
ws.append(['序号','刀具信息','刀具长度']) #增加标题栏
ws.row_dimensions[1].height = 30

for m in range(0,len(final)):      
    print(final[m][0])      
    d = m+1,final[m][0],final[m][1]   #显示序号，刀具信息，刀具长度         
    ws.append(d)   
    ws.row_dimensions[m + 2].height = 30 #设置单元格高度

ws.column_dimensions['A'].width = 10 #设置单元格宽度
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15

border=Border(top=Side(border_style='thin',color=colors.BLACK), #设置单元格线形，颜色
              bottom=Side(border_style='thin',color=colors.BLACK),
              left=Side(border_style='thin',color=colors.BLACK),
              right=Side(border_style='thin',color=colors.BLACK))

for i in range(1,ws.max_row+1):
    for j in range(1,ws.max_column+1):
        ws.cell(row=i, column=j).border=border  #设置单元格的对齐方式
        ws.cell(row=i+1, column=j).alignment=Alignment(horizontal='left',vertical='center')
        ws.cell(row=1, column=j).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=i, column=1).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=i, column=j).font=Font(size=15)
 
#print(ws) 
wb.save('tool.xlsx')   
wb.close()          
print('成功写入')
print('程序将在3s内自动关闭..')
for i in range(3,0,-1):    
    time.sleep(1) 
    print(i)
    
print('Thank you !!!')
#print(result)

 
 